import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 1. CONFIGURAÇÃO DA PÁGINA E CSS
st.set_page_config(page_title="Conciliador PRO | HITS x Getnet", layout="wide", page_icon="📈")

st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
    :root { --primary: #11CAA0; --dark-navy: #002c51; --light-bg: #f8fafc; --card-bg: #ffffff; }
    .stApp { background-color: var(--light-bg); font-family: 'Inter', sans-serif; }
    h1 { color: var(--dark-navy) !important; font-weight: 700 !important; }
    p { color: #64748b !important; }
    .stFileUploader {
        border: 2px dashed var(--primary) !important; border-radius: 15px !important;
        background-color: var(--card-bg) !important; padding: 20px !important; transition: transform 0.3s ease;
    }
    .stButton>button {
        width: 100% !important; background: linear-gradient(135deg, #11CAA0 0%, #0da582 100%) !important;
        border-radius: 10px !important; border: none !important; padding: 15px !important; transition: 0.3s all !important;
    }
    .stButton>button div p, .stButton>button span, .stButton>button {
        color: white !important; font-weight: 700 !important; font-size: 16px !important;
    }
    div[data-testid="metric-container"] {
        background: white; padding: 15px; border-radius: 12px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border-bottom: 4px solid var(--primary);
    }
    [data-testid="stElementToolbar"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)

# --- FUNÇÕES DE PROCESSAMENTO E MATEMÁTICAS ---

def garantir_numero(serie):
    if serie.dtype == 'object':
        serie = serie.astype(str).str.replace('R$', '', regex=False).str.strip()
        serie = serie.str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
    return pd.to_numeric(serie, errors='coerce').fillna(0)

def limpar_cv(valor):
    v = str(valor).strip().upper()
    if v in ['NAN', 'NONE', 'NAT', 'NULL', '']: return ''
    if v.endswith('.0'): v = v[:-2]
    return v

def levenshtein(s1, s2):
    """Calcula a distância de edições entre duas strings (Tolerância de digitação)"""
    if len(s1) < len(s2): return levenshtein(s2, s1)
    if len(s2) == 0: return len(s1)
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    return previous_row[-1]

def ler_excel_inteligente(file, palavra_chave, aba=0):
    try:
        df_temp = pd.read_excel(file, header=None, nrows=25, sheet_name=aba)
        for indice, linha in df_temp.iterrows():
            if linha.astype(str).str.contains(palavra_chave, case=False, na=False).any():
                return pd.read_excel(file, header=indice, sheet_name=aba)
    except: return pd.DataFrame()
    return pd.read_excel(file, sheet_name=aba)

def formata_moeda(val):
    if pd.isna(val) or val == '': return ''
    try: return f"R$ {float(val):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except: return val

def formatar_usuario(nome_cru):
    nome_str = str(nome_cru).strip()
    if not nome_str or nome_str.lower() in ['nan', 'none', 'nat', '<na>']: return ''
    match = re.search(r'([^,]+),\s*HOTEL\s*-\s*(.+)', nome_str, re.IGNORECASE)
    if match:
        sobrenome = match.group(1).strip().title()
        nome = match.group(2).strip().title()
        return f"{nome} {sobrenome}"
    return nome_str.title()

def obter_categoria_macro(mod_string):
    m = str(mod_string).upper().strip()
    tipo = ""
    if "CRED" in m or "CRÉD" in m: tipo = "CRÉDITO"
    elif "DEB" in m or "DÉB" in m: tipo = "DÉBITO"
    elif "PIX" in m: return "PIX"
    
    bandeira = "OUTROS"
    if "MASTER" in m: bandeira = "MASTERCARD"
    elif "VISA" in m: bandeira = "VISA"
    elif "ELO" in m: bandeira = "ELO"
    elif "AMEX" in m or "AMERICAN" in m: bandeira = "AMEX"
    elif "HIPER" in m: bandeira = "HIPERCARD"
    
    if tipo: return f"{bandeira} {tipo}"
    return m if m else "NÃO IDENTIFICADO"

# --- INTERFACE E LÓGICA PRINCIPAL ---

st.markdown("<h1 style='text-align: center;'>Conciliação Financeira HITS x Getnet</h1>", unsafe_allow_html=True)
st.markdown("<p style='text-align: center; margin-bottom: 40px;'>Arraste seus relatórios abaixo para iniciar o cruzamento inteligente.</p>", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1: hits_file = st.file_uploader("🏨 Relatório HITS", type=["xlsx"], key="hits")
with col2: getnet_file = st.file_uploader("💳 Relatório Getnet", type=["xlsx"], key="getnet")

if hits_file and getnet_file:
    if st.button("ANALISAR E CONCILIAR AGORA"):
        with st.spinner("Processando Inteligência Financeira..."):
            
            # --- 1. GETNET ---
            df_g_cartoes = ler_excel_inteligente(getnet_file, 'BANDEIRA', aba=0)
            df_g_cartoes.columns = df_g_cartoes.columns.astype(str).str.strip().str.upper()
            
            if 'MODALIDADE' not in df_g_cartoes.columns:
                st.error("❌ ERRO: Coluna de Modalidade ausente. Verifique se os arquivos não foram invertidos.")
                st.stop()

            if 'STATUS DA TRANSAÇÃO' in df_g_cartoes.columns:
                df_g_cartoes = df_g_cartoes[df_g_cartoes['STATUS DA TRANSAÇÃO'].str.contains('Aprovada', case=False, na=False)]
            
            # Busca coluna de parcelas para implementar a regra de sufixo
            col_parcelas = next((c for c in df_g_cartoes.columns if 'PARCELA' in str(c)), None)
            
            df_g_cartoes = df_g_cartoes.rename(columns={
                'NÚMERO DE AUTORIZAÇÃO (AUT)': 'Auto_G', 'NÚMERO DO COMPROVANTE DE VENDAS (CV)': 'CV_G',
                'VALOR BRUTO': 'Valor_G', 'DATA/HORA DA VENDA': 'Data_G', 'MODALIDADE': 'Mod_G', 'BANDEIRA': 'Band_G'
            })
            df_g_cartoes = df_g_cartoes[~df_g_cartoes['Mod_G'].astype(str).str.upper().str.contains('GET ECO', na=False)]
            
            # Aplica sufixo de parcelas se houver
            def formatar_mod_getnet(row):
                base_mod = f"{row['Band_G']} {row['Mod_G']}"
                if col_parcelas and pd.notna(row[col_parcelas]):
                    parc = str(row[col_parcelas]).replace('.0', '').strip()
                    if parc.isdigit() and int(parc) > 1:
                        return f"{base_mod} - {parc}x"
                return base_mod

            df_g_cartoes['Modalidade_G'] = df_g_cartoes.apply(formatar_mod_getnet, axis=1)

            df_g_pix = ler_excel_inteligente(getnet_file, 'VALOR', aba='PIX')
            if not df_g_pix.empty:
                df_g_pix.columns = df_g_pix.columns.astype(str).str.strip().str.upper()
                col_st_pix = next((c for c in df_g_pix.columns if 'STATUS' in str(c)), None)
                if col_st_pix: df_g_pix = df_g_pix[df_g_pix[col_st_pix].astype(str).str.contains('Paga', case=False, na=False)]
                col_v_pix = next((c for c in df_g_pix.columns if 'VALOR' in str(c)), None)
                col_d_pix = next((c for c in df_g_pix.columns if 'DATA' in str(c)), None)
                df_g_pix = pd.DataFrame({
                    'Valor_G': garantir_numero(df_g_pix[col_v_pix]) if col_v_pix else 0,
                    'Data_G': df_g_pix[col_d_pix] if col_d_pix else '',
                    'Modalidade_G': 'GETNET PIX', 'Auto_G': 'PIX_SEM_AUT', 'CV_G': ''
                })

            # --- 2. HITS E EXTRAÇÃO DE DINHEIRO ---
            df_hits = ler_excel_inteligente(hits_file, 'Autorização')
            df_hits.columns = df_hits.columns.astype(str).str.strip()
            
            if 'Usuário' not in df_hits.columns: df_hits['Usuário'] = ''
            if 'Conta' not in df_hits.columns: df_hits['Conta'] = ''
                
            df_hits = df_hits.rename(columns={
                'Autorização': 'Auto_H', 'Documento': 'CV_H', 'Valor': 'Valor_H', 
                'Data': 'Data_H', 'Pagamento': 'Pagamento', 'Tipo de Pagamento': 'Modalidade_H', 
                'Usuário': 'Usuário', 'Conta': 'Conta'
            })
            
            df_hits['Usuário'] = df_hits['Usuário'].apply(formatar_usuario)

            # EXTRAÇÃO DO DINHEIRO
            mask_dinheiro = df_hits['Modalidade_H'].astype(str).str.upper().str.contains('DINHEIRO', na=False)
            df_dinheiro = df_hits[mask_dinheiro].copy()
            
            df_dinheiro['Valor_H'] = garantir_numero(df_dinheiro['Valor_H'])
            df_dinheiro['Data_H'] = pd.to_datetime(df_dinheiro['Data_H'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
            df_dinheiro_resumo = df_dinheiro.groupby(['Data_H', 'Usuário'], as_index=False)['Valor_H'].sum()
            df_dinheiro_resumo.rename(columns={'Data_H': 'Data', 'Valor_H': 'Total Recebido'}, inplace=True)
            df_dinheiro_resumo = df_dinheiro_resumo.sort_values(by=['Data', 'Usuário'])
            
            filtro_h = 'FATURADO|DINHEIRO|GET ECO|CENTRAL TRANSFERENCIA/PIX'
            df_hits = df_hits[~df_hits['Modalidade_H'].astype(str).str.upper().str.contains(filtro_h, regex=True)]

            # CRIAÇÃO DA TABELA COMPARATIVA MACRO POR BANDEIRA / MODALIDADE
            df_h_macro = df_hits.copy()
            df_h_macro['Categoria'] = df_h_macro['Modalidade_H'].apply(obter_categoria_macro)
            res_h = df_h_macro.groupby('Categoria')['Valor_H'].apply(lambda x: garantir_numero(x).sum()).reset_index()
            res_h.rename(columns={'Valor_H': 'Total HITS'}, inplace=True)
            
            g_list = []
            if not df_g_cartoes.empty:
                g_list.append(df_g_cartoes[['Modalidade_G', 'Valor_G']].rename(columns={'Modalidade_G': 'Mod', 'Valor_G': 'Val'}))
            if not df_g_pix.empty:
                g_list.append(df_g_pix[['Modalidade_G', 'Valor_G']].rename(columns={'Modalidade_G': 'Mod', 'Valor_G': 'Val'}))
            
            if g_list:
                df_g_macro = pd.concat(g_list, ignore_index=True)
                df_g_macro['Categoria'] = df_g_macro['Mod'].apply(obter_categoria_macro)
                res_g = df_g_macro.groupby('Categoria')['Val'].sum().reset_index()
                res_g.rename(columns={'Val': 'Total Getnet'}, inplace=True)
            else:
                res_g = pd.DataFrame(columns=['Categoria', 'Total Getnet'])
                
            df_macro_resumo = pd.merge(res_h, res_g, on='Categoria', how='outer').fillna(0)
            df_macro_resumo['Diferença'] = df_macro_resumo['Total HITS'] - df_macro_resumo['Total Getnet']

            # Separar o PIX MANUAL
            mask_manual = df_hits['Modalidade_H'].astype(str).str.upper().str.contains('MANUAL', na=False)
            df_hits_manual = df_hits[mask_manual].copy()
            df_hits = df_hits[~mask_manual].copy()

            # --- 3. CRUZAMENTOS MAIN ---
            mask_pix_h = df_hits['Modalidade_H'].astype(str).str.upper().str.contains('PIX', na=False)
            df_h_pix, df_h_cart = df_hits[mask_pix_h].copy(), df_hits[~mask_pix_h].copy()

            for df in [df_h_cart, df_g_cartoes]:
                col_a = 'Auto_H' if 'Auto_H' in df.columns else 'Auto_G'
                df[col_a] = df[col_a].astype(str).str.strip().str.upper()
                df['Valor_H' if 'Valor_H' in df.columns else 'Valor_G'] = garantir_numero(df['Valor_H' if 'Valor_H' in df.columns else 'Valor_G'])

            invalid_autos = ['', '0', '0.0', 'NAN', 'NONE', 'NULL', 'X', '*', 'PIX_SEM_AUT']
            
            mask_h_valid = ~df_h_cart['Auto_H'].isin(invalid_autos) & ~df_h_cart['Auto_H'].str.match(r'^0+$', na=True)
            mask_g_valid = ~df_g_cartoes['Auto_G'].isin(invalid_autos) & ~df_g_cartoes['Auto_G'].str.match(r'^0+$', na=True)
            
            df_h_valid = df_h_cart[mask_h_valid].copy()
            df_h_invalid = df_h_cart[~mask_h_valid].copy()
            
            df_g_valid = df_g_cartoes[mask_g_valid].copy()
            df_g_invalid = df_g_cartoes[~mask_g_valid].copy()

            df_h_valid['Card_Match'] = df_h_valid.groupby('Auto_H').cumcount()
            df_g_valid['Card_Match'] = df_g_valid.groupby('Auto_G').cumcount()

            # Merge exato inicial pelos Autos
            df_m_cart_valid = pd.merge(
                df_h_valid, 
                df_g_valid[['Auto_G', 'CV_G', 'Valor_G', 'Data_G', 'Modalidade_G', 'Card_Match']], 
                left_on=['Auto_H', 'Card_Match'], 
                right_on=['Auto_G', 'Card_Match'], 
                how='outer', 
                indicator=True
            ).drop(columns=['Card_Match'])

            df_h_invalid['_merge'] = 'left_only'
            df_g_invalid['_merge'] = 'right_only'
            
            df_m_cart = pd.concat([df_m_cart_valid, df_h_invalid, df_g_invalid], ignore_index=True)

            if not df_g_pix.empty:
                df_h_pix['Valor_H'] = garantir_numero(df_h_pix['Valor_H'])
                df_g_pix['Valor_G'] = garantir_numero(df_g_pix['Valor_G'])
                
                df_h_pix['Dt_Datetime'] = pd.to_datetime(df_h_pix['Data_H'], errors='coerce', dayfirst=True)
                df_g_pix['Dt_Datetime'] = pd.to_datetime(df_g_pix['Data_G'], errors='coerce', dayfirst=True)
                
                df_h_pix['Dt_Day'] = df_h_pix['Dt_Datetime'].dt.date
                df_g_pix['Dt_Day'] = df_g_pix['Dt_Datetime'].dt.date
                
                df_h_pix = df_h_pix.sort_values('Dt_Datetime')
                df_g_pix = df_g_pix.sort_values('Dt_Datetime')
                
                df_h_pix['Match'] = df_h_pix.groupby(['Dt_Day', df_h_pix['Valor_H'].round(2)]).cumcount()
                df_g_pix['Match'] = df_g_pix.groupby(['Dt_Day', df_g_pix['Valor_G'].round(2)]).cumcount()
                
                df_m_pix = pd.merge(
                    df_h_pix, df_g_pix, 
                    left_on=['Dt_Day', 'Valor_H', 'Match'], 
                    right_on=['Dt_Day', 'Valor_G', 'Match'], 
                    how='outer', indicator=True
                ).drop(columns=['Match', 'Dt_Day', 'Dt_Datetime_x', 'Dt_Datetime_y'], errors='ignore')
            else:
                df_h_pix['_merge'] = 'left_only'
                df_m_pix = df_h_pix

            # --- 4. TRATAMENTO, PAREAMENTO SECUNDÁRIO E STATUS ---
            df_res = pd.concat([df_m_cart, df_m_pix], ignore_index=True)
            df_res['ID'] = '' 
            
            df_res['CV_H'] = df_res['CV_H'].apply(limpar_cv)
            df_res['CV_G'] = df_res['CV_G'].apply(limpar_cv)
            
            df_res['Data_H'] = pd.to_datetime(df_res['Data_H'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
            df_res['Data_G'] = pd.to_datetime(df_res['Data_G'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
            df_res['Auto_H'] = df_res['Auto_H'].fillna('').astype(str).str.strip().str.upper()
            df_res['Auto_G'] = df_res['Auto_G'].fillna('').astype(str).str.strip().str.upper()

            df_res['Status'] = 'Falta no HITS'
            df_res.loc[df_res['_merge'] == 'left_only', 'Status'] = 'Falta na Getnet'
            
            if not df_hits_manual.empty:
                df_hits_manual['ID'] = ''
                df_hits_manual['Status'] = 'A VERIFICAR'
                df_hits_manual['_merge'] = 'left_only'
                df_hits_manual['Data_H'] = pd.to_datetime(df_hits_manual['Data_H'], errors='coerce', dayfirst=True).dt.strftime('%d/%m/%Y')
                df_res = pd.concat([df_res, df_hits_manual], ignore_index=True)

            id_count = 1

            # --- PAREAMENTO SECUNDÁRIO (Para pegar os erros de digitação de Auto) ---
            mask_fh2 = (df_res['Status'] == 'Falta na Getnet') & (df_res['ID'] == '')
            mask_fg2 = (df_res['Status'] == 'Falta no HITS') & (df_res['ID'] == '')
            
            for idx_h in df_res[mask_fh2].index:
                for idx_g in df_res[mask_fg2].index:
                    if df_res.loc[idx_g, 'ID'] != '':
                        continue
                    v_h = pd.to_numeric(df_res.loc[idx_h, 'Valor_H'], errors='coerce') or 0
                    v_g = pd.to_numeric(df_res.loc[idx_g, 'Valor_G'], errors='coerce') or 0
                    
                    # Checagem primária de valor idêntico e data idêntica (para habilitar o match de sobras)
                    if np.isclose(v_h, v_g, atol=0.001):
                        dt_h = str(df_res.loc[idx_h, 'Data_H']).strip()
                        dt_g = str(df_res.loc[idx_g, 'Data_G']).strip()
                        cv_h = str(df_res.loc[idx_h, 'CV_H']).strip()
                        cv_g = str(df_res.loc[idx_g, 'CV_G']).strip()
                        
                        if (dt_h == dt_g) or (cv_h == cv_g and cv_h != ''):
                            df_res.loc[idx_h, 'ID'] = df_res.loc[idx_g, 'ID'] = f'#{id_count}'
                            
                            # Transfere os dados da Getnet para a linha do HITS para unificar visualmente
                            df_res.loc[idx_h, 'Valor_G'] = df_res.loc[idx_g, 'Valor_G']
                            df_res.loc[idx_h, 'Data_G'] = df_res.loc[idx_g, 'Data_G']
                            df_res.loc[idx_h, 'Modalidade_G'] = df_res.loc[idx_g, 'Modalidade_G']
                            df_res.loc[idx_h, 'Auto_G'] = df_res.loc[idx_g, 'Auto_G']
                            df_res.loc[idx_h, 'CV_G'] = df_res.loc[idx_g, 'CV_G']
                            df_res.loc[idx_h, '_merge'] = 'both'
                            
                            df_res = df_res.drop(idx_g) # Remove a sobra que foi unificada
                            id_count += 1
                            break

            # --- PROCESSAMENTO DEFINITIVO DE STATUS (REGRAS CLARAS DE PRIORIDADE E FUZZY MATCH) ---
            for idx in df_res[df_res['_merge'] == 'both'].index:
                v_h = pd.to_numeric(df_res.loc[idx, 'Valor_H'], errors='coerce') or 0
                v_g = pd.to_numeric(df_res.loc[idx, 'Valor_G'], errors='coerce') or 0
                cv_h = str(df_res.loc[idx, 'CV_H']).strip()
                cv_g = str(df_res.loc[idx, 'CV_G']).strip()
                auto_h = str(df_res.loc[idx, 'Auto_H']).strip()
                auto_g = str(df_res.loc[idx, 'Auto_G']).strip()
                dt_h = str(df_res.loc[idx, 'Data_H']).strip()
                dt_g = str(df_res.loc[idx, 'Data_G']).strip()
                
                # Respeita a modalidade real (sem atalhos perigosos)
                mod_h_macro = obter_categoria_macro(df_res.loc[idx, 'Modalidade_H'])
                mod_g_macro = obter_categoria_macro(df_res.loc[idx, 'Modalidade_G'])
                
                is_pix = 'PIX' in mod_h_macro or 'PIX' in mod_g_macro

                # Validação Fuzzy (Tolerância Matemática)
                dist_auto = levenshtein(auto_h, auto_g)
                dist_cv = levenshtein(cv_h, cv_g)
                
                auto_match = (auto_h == auto_g) or (dist_auto <= 2 and auto_h != '' and auto_g != '')
                cv_match = (cv_h == cv_g) or (dist_cv <= 2 and cv_h != '' and cv_g != '')

                # PRIORIDADE DE JULGAMENTO
                if not np.isclose(v_h, v_g, atol=0.001):
                    df_res.loc[idx, 'Status'] = 'VALOR INCORRETO'
                elif mod_h_macro != mod_g_macro:
                    df_res.loc[idx, 'Status'] = 'ERRO DE MODALIDADE'
                elif dt_h != dt_g:
                    df_res.loc[idx, 'Status'] = 'DATA INCORRETA'
                elif is_pix:
                    df_res.loc[idx, 'Status'] = 'Batido - OK' # Pix isento de regras de Auto/CV
                else:
                    if auto_match and cv_match:
                        df_res.loc[idx, 'Status'] = 'Batido - OK'
                    elif auto_h == '' and cv_h == '':
                        df_res.loc[idx, 'Status'] = 'AUTO/CV AUSENTES'
                    elif not auto_match and not cv_match:
                        df_res.loc[idx, 'Status'] = 'AUTO/CV INCORRETOS'
                    elif not auto_match:
                        df_res.loc[idx, 'Status'] = 'AUTO INCORRETO'
                    elif not cv_match:
                        df_res.loc[idx, 'Status'] = 'CV INCORRETO'

            # --- DETECTOR DE AUTO/CV TROCADOS ENTRE LANÇAMENTOS DIFERENTES ---
            indices_both = df_res[df_res['_merge'] == 'both'].index.tolist()
            used_trocados = set()
            
            for idx_i in indices_both:
                if idx_i in used_trocados or df_res.loc[idx_i, 'Status'] == 'Batido - OK':
                    continue
                v_h_i = pd.to_numeric(df_res.loc[idx_i, 'Valor_H'], errors='coerce') or 0
                v_g_i = pd.to_numeric(df_res.loc[idx_i, 'Valor_G'], errors='coerce') or 0
                
                for idx_j in indices_both:
                    if idx_j == idx_i or idx_j in used_trocados or df_res.loc[idx_j, 'Status'] == 'Batido - OK':
                        continue
                    v_h_j = pd.to_numeric(df_res.loc[idx_j, 'Valor_H'], errors='coerce') or 0
                    v_g_j = pd.to_numeric(df_res.loc[idx_j, 'Valor_G'], errors='coerce') or 0
                    
                    if np.isclose(v_h_i, v_g_j, atol=0.001) and np.isclose(v_g_i, v_h_j, atol=0.001):
                        df_res.loc[idx_i, 'Status'] = 'AUTO/CV TROCADOS'
                        df_res.loc[idx_j, 'Status'] = 'AUTO/CV TROCADOS'
                        if df_res.loc[idx_i, 'ID'] == '': df_res.loc[idx_i, 'ID'] = f'#{id_count}'
                        if df_res.loc[idx_j, 'ID'] == '': df_res.loc[idx_j, 'ID'] = f'#{id_count}'
                        id_count += 1
                        used_trocados.add(idx_i)
                        used_trocados.add(idx_j)
                        break

            # --- DETECTOR MATRICIAL DE PAGAMENTOS DUPLICADOS ---
            autos_com_getnet = set(df_res[df_res['_merge'] == 'both']['Auto_H'].astype(str).str.strip().str.upper())
            for p in invalid_autos: autos_com_getnet.discard(p)

            pix_casados = set()
            df_both_pix = df_res[df_res['_merge'] == 'both']
            for idx_b in df_both_pix.index:
                if 'PIX' in str(df_both_pix.loc[idx_b, 'Modalidade_H']).upper():
                    dt_val = str(df_both_pix.loc[idx_b, 'Data_H']).strip()
                    try:
                        v_val = float(df_both_pix.loc[idx_b, 'Valor_H'])
                        pix_casados.add((dt_val, round(v_val, 2)))
                    except: pass

            mask_falta_g = df_res['Status'] == 'Falta na Getnet'
            for idx in df_res[mask_falta_g].index:
                auto_atual = str(df_res.loc[idx, 'Auto_H']).strip().upper()
                mod_atual = str(df_res.loc[idx, 'Modalidade_H']).upper()
                
                if auto_atual in autos_com_getnet and auto_atual != '':
                    df_res.loc[idx, 'Status'] = 'PAGAMENTO DUPLICADO'
                elif 'PIX' in mod_atual:
                    dt_atual = str(df_res.loc[idx, 'Data_H']).strip()
                    try:
                        v_atual = float(df_res.loc[idx, 'Valor_H'])
                        if (dt_atual, round(v_atual, 2)) in pix_casados:
                            df_res.loc[idx, 'Status'] = 'PAGAMENTO DUPLICADO'
                    except: pass

            # --- IDENTIFICAR AUTO/CV INVERTIDO NA MESMA LINHA ---
            mask_err = df_res['Status'].isin(['AUTO INCORRETO', 'CV INCORRETO', 'AUTO/CV INCORRETOS'])
            for idx in df_res[mask_err].index:
                auto_h = str(df_res.loc[idx, 'Auto_H']).strip()
                cv_h = str(df_res.loc[idx, 'CV_H']).strip()
                auto_g = str(df_res.loc[idx, 'Auto_G']).strip()
                cv_g = str(df_res.loc[idx, 'CV_G']).strip()
                if auto_h == cv_g and cv_h == auto_g and auto_h != '' and cv_h != '':
                    df_res.loc[idx, 'Status'] = 'AUTO/CV INVERTIDO'

            # --- ATRIBUIÇÃO FINAL DE IDs E ORDENAÇÃO ---
            mask_erros_geral = df_res['Status'].isin(['CV INCORRETO', 'ERRO DE MODALIDADE', 'DATA INCORRETA', 'VALOR INCORRETO', 'AUTO INCORRETO', 'AUTO/CV INCORRETOS', 'AUTO/CV AUSENTES', 'AUTO/CV INVERTIDO', 'AUTO/CV TROCADOS', 'PAGAMENTO DUPLICADO', 'Falta na Getnet', 'Falta no HITS', 'A VERIFICAR'])
            for idx in df_res[mask_erros_geral].index:
                if df_res.loc[idx, 'ID'] == '':
                    df_res.loc[idx, 'ID'] = f'#{id_count}'
                    id_count += 1

            mapa_ordem = {'Falta na Getnet':1, 'Falta no HITS':2, 'PAGAMENTO DUPLICADO':3, 'AUTO/CV TROCADOS':4, 'AUTO/CV INVERTIDO':5, 'AUTO/CV INCORRETOS':6, 'AUTO/CV AUSENTES':7, 'AUTO INCORRETO':8, 'CV INCORRETO':9, 'ERRO DE MODALIDADE':10, 'DATA INCORRETA':11, 'VALOR INCORRETO':12, 'A VERIFICAR':13, 'Batido - OK':14}
            df_res['Ordem'] = df_res['Status'].map(mapa_ordem).fillna(99)
            df_res = df_res.sort_values(by=['Ordem', 'ID', 'Data_H']).reset_index(drop=True)
            
            # --- RASTREIO DE IDs PARA A ABA MACRO ---
            df_res['Categoria_Macro'] = df_res['Modalidade_H'].apply(obter_categoria_macro)
            
            rastreio_ids = {}
            for idx in df_res.index:
                cat = df_res.loc[idx, 'Categoria_Macro']
                if df_res.loc[idx, 'Status'] != 'Batido - OK':
                    id_val = str(df_res.loc[idx, 'ID']).strip()
                    if id_val:
                        if cat not in rastreio_ids:
                            rastreio_ids[cat] = set()
                        rastreio_ids[cat].add(id_val)
            
            df_macro_resumo['Lançamentos Causadores (IDs)'] = df_macro_resumo['Categoria'].map(
                lambda c: ", ".join(sorted(rastreio_ids.get(c, [])))
            )

            cols_f = ['ID', 'Status', 'Pagamento', 'Conta', 'Valor_H', 'Valor_G', 'Auto_H', 'Auto_G', 'CV_H', 'CV_G', 'Data_H', 'Data_G', 'Modalidade_H', 'Modalidade_G', 'Usuário']
            df_res = df_res[[c for c in cols_f if c in df_res.columns]].fillna('')
            for c in df_res.columns: df_res[c] = df_res[c].apply(lambda x: '' if str(x).strip().lower() in ['none', 'nan', 'nat', '<na>'] else x)

            # --- PINTURA CIRÚRGICA (WEB INTERFACE) ---
            def cor_tela(row):
                est = [''] * len(row)
                cols = list(row.index)
                st_val = row['Status']
                
                if st_val == 'Batido - OK': est = ['background-color: #e6ffed'] * len(row)
                elif st_val == 'Falta na Getnet':
                    for c in ['Pagamento', 'Conta', 'Valor_H', 'Auto_H', 'CV_H', 'Data_H', 'Modalidade_H', 'Usuário']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffeef0'
                elif st_val == 'Falta no HITS':
                    for c in ['Valor_G', 'CV_G', 'Auto_G', 'Data_G', 'Modalidade_G']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffeef0'
                elif st_val == 'A VERIFICAR':
                    if 'Status' in cols: est[cols.index('Status')] = 'background-color: #d0ebff; font-weight: bold; color: #004085;'
                
                elif st_val == 'CV INCORRETO':
                    if 'CV_H' in cols and row['CV_H'] != '': est[cols.index('CV_H')] = 'background-color: #ffb067; font-weight: bold;'
                    if 'CV_G' in cols and row['CV_G'] != '': est[cols.index('CV_G')] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'ERRO DE MODALIDADE':
                    if 'Modalidade_H' in cols and row['Modalidade_H'] != '': est[cols.index('Modalidade_H')] = 'background-color: #ffb067; font-weight: bold;'
                    if 'Modalidade_G' in cols and row['Modalidade_G'] != '': est[cols.index('Modalidade_G')] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'DATA INCORRETA':
                    if 'Data_H' in cols and row['Data_H'] != '': est[cols.index('Data_H')] = 'background-color: #ffb067; font-weight: bold;'
                    if 'Data_G' in cols and row['Data_G'] != '': est[cols.index('Data_G')] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'VALOR INCORRETO':
                    if 'Valor_H' in cols and row['Valor_H'] != '': est[cols.index('Valor_H')] = 'background-color: #ffb067; font-weight: bold;'
                    if 'Valor_G' in cols and row['Valor_G'] != '': est[cols.index('Valor_G')] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'AUTO INCORRETO':
                    if 'Auto_H' in cols and row['Auto_H'] != '': est[cols.index('Auto_H')] = 'background-color: #ffb067; font-weight: bold;'
                    if 'Auto_G' in cols and row['Auto_G'] != '': est[cols.index('Auto_G')] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'AUTO/CV INCORRETOS':
                    for c in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'AUTO/CV AUSENTES':
                    for c in ['Auto_H', 'CV_H']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'AUTO/CV INVERTIDO':
                    for c in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'AUTO/CV TROCADOS':
                    for c in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G', 'Valor_H', 'Valor_G']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffb067; font-weight: bold;'
                elif st_val == 'PAGAMENTO DUPLICADO':
                    for c in ['Pagamento', 'Conta', 'Valor_H', 'Auto_H', 'CV_H', 'Data_H', 'Modalidade_H', 'Usuário']:
                        if c in cols: est[cols.index(c)] = 'background-color: #ffb067; font-weight: bold;'
                
                if str(row.get('ID', '')).strip() != '' and 'ID' in cols:
                    est[cols.index('ID')] = 'background-color: #fce83a; font-weight: bold; color: black;'
                    
                return est

            # --- DASHBOARD ---
            st.success("✅ Conciliação Realizada!")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Total", len(df_res))
            c2.metric("OK", len(df_res[df_res['Status'] == 'Batido - OK']))
            c3.metric("Faltas", len(df_res[df_res['Status'].str.contains('Falta')]))
            c4.metric("Inconsistências", len(df_res[~df_res['Status'].isin(['Batido - OK', 'A VERIFICAR']) & ~df_res['Status'].str.contains('Falta')]))
            c5.metric("A Verificar", len(df_res[df_res['Status'] == 'A VERIFICAR']))

            st.dataframe(df_res.style.apply(cor_tela, axis=1).format({'Valor_H': formata_moeda, 'Valor_G': formata_moeda}), use_container_width=True)
            
            st.markdown("### 📊 Volume Macro por Bandeira/Modalidade")
            st.dataframe(df_macro_resumo.style.format({'Total HITS': formata_moeda, 'Total Getnet': formata_moeda, 'Diferença': formata_moeda}), use_container_width=True)

            # --- EXPORTAÇÃO EXCEL PROFISSIONAL ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                # PLANILHA 1: RESULTADO PRINCIPAL
                df_res.to_excel(writer, index=False, sheet_name='Resultado')
                ws = writer.sheets['Resultado']
                
                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions
                
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[col_letter].width = min((max_length + 2), 35)

                f_ok, f_red, f_org, f_blu, f_ylw = PatternFill("solid", "E6FFED"), PatternFill("solid", "FFEEF0"), PatternFill("solid", "FFB067"), PatternFill("solid", "D0EBFF"), PatternFill("solid", "FCE83A")
                center_align = Alignment(horizontal="center", vertical="center")
                
                for c in range(1, ws.max_column + 1):
                    ws.cell(1, c).alignment = center_align

                idx = {n: i for i, n in enumerate(df_res.columns, 1)}

                for r in range(2, ws.max_row + 1):
                    st_v = ws.cell(r, idx['Status']).value
                    id_val = str(ws.cell(r, idx['ID']).value or '').strip()
                    
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).alignment = center_align
                    
                    for c_n in ['Valor_H', 'Valor_G']:
                        if c_n in idx and ws.cell(r, idx[c_n]).value != '':
                            ws.cell(r, idx[c_n]).number_format = '"R$" #,##0.00'
                    
                    # Pintura Excel Espelhada com as Novas Regras
                    if st_v == 'Batido - OK':
                        for c in range(1, ws.max_column + 1): ws.cell(r, c).fill = f_ok
                    elif st_v == 'Falta na Getnet':
                        for c_n in ['Pagamento', 'Conta', 'Valor_H', 'Auto_H', 'CV_H', 'Data_H', 'Modalidade_H', 'Usuário']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_red
                    elif st_v == 'Falta no HITS':
                        for c_n in ['Valor_G', 'CV_G', 'Auto_G', 'Data_G', 'Modalidade_G']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_red
                    elif st_v == 'A VERIFICAR':
                        ws.cell(r, idx['Status']).fill = f_blu
                    
                    elif st_v == 'CV INCORRETO':
                        if 'CV_H' in idx and ws.cell(r, idx['CV_H']).value != '': ws.cell(r, idx['CV_H']).fill = f_org
                        if 'CV_G' in idx and ws.cell(r, idx['CV_G']).value != '': ws.cell(r, idx['CV_G']).fill = f_org
                    elif st_v == 'ERRO DE MODALIDADE':
                        if 'Modalidade_H' in idx and ws.cell(r, idx['Modalidade_H']).value != '': ws.cell(r, idx['Modalidade_H']).fill = f_org
                        if 'Modalidade_G' in idx and ws.cell(r, idx['Modalidade_G']).value != '': ws.cell(r, idx['Modalidade_G']).fill = f_org
                    elif st_v == 'DATA INCORRETA':
                        if 'Data_H' in idx and ws.cell(r, idx['Data_H']).value != '': ws.cell(r, idx['Data_H']).fill = f_org
                        if 'Data_G' in idx and ws.cell(r, idx['Data_G']).value != '': ws.cell(r, idx['Data_G']).fill = f_org
                    elif st_v == 'VALOR INCORRETO':
                        if 'Valor_H' in idx and ws.cell(r, idx['Valor_H']).value != '': ws.cell(r, idx['Valor_H']).fill = f_org
                        if 'Valor_G' in idx and ws.cell(r, idx['Valor_G']).value != '': ws.cell(r, idx['Valor_G']).fill = f_org
                    elif st_v == 'AUTO INCORRETO':
                        if 'Auto_H' in idx and ws.cell(r, idx['Auto_H']).value != '': ws.cell(r, idx['Auto_H']).fill = f_org
                        if 'Auto_G' in idx and ws.cell(r, idx['Auto_G']).value != '': ws.cell(r, idx['Auto_G']).fill = f_org
                    elif st_v == 'AUTO/CV INCORRETOS':
                        for c_n in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_org
                    elif st_v == 'AUTO/CV AUSENTES':
                        for c_n in ['Auto_H', 'CV_H']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_org
                    elif st_v == 'AUTO/CV INVERTIDO':
                        for c_n in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_org
                    elif st_v == 'AUTO/CV TROCADOS':
                        for c_n in ['Auto_H', 'Auto_G', 'CV_H', 'CV_G', 'Valor_H', 'Valor_G']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_org
                    elif st_v == 'PAGAMENTO DUPLICADO':
                        for c_n in ['Pagamento', 'Conta', 'Valor_H', 'Auto_H', 'CV_H', 'Data_H', 'Modalidade_H', 'Usuário']:
                            if c_n in idx: ws.cell(r, idx[c_n]).fill = f_org

                    if id_val != '' and 'ID' in idx:
                        ws.cell(r, idx['ID']).fill = f_ylw
                
                # PLANILHA 2: RESUMO DINHEIRO INTELIGENTE
                if not df_dinheiro_resumo.empty:
                    df_dinheiro_resumo.to_excel(writer, index=False, sheet_name='Dinheiro', startcol=0)
                    df_dinheiro_totais = df_dinheiro_resumo.groupby('Data', as_index=False)['Total Recebido'].sum()
                    df_dinheiro_totais.rename(columns={'Total Recebido': 'Total do Dia'}, inplace=True)
                    df_dinheiro_totais.to_excel(writer, index=False, sheet_name='Dinheiro', startcol=4)
                    
                    ws_din = writer.sheets['Dinheiro']
                    ws_din.freeze_panes = 'A2'
                    
                    for column in ws_din.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        ws_din.column_dimensions[col_letter].width = min((max_length + 2), 35)
                        
                    for r in range(1, ws_din.max_row + 1):
                        for c in range(1, ws_din.max_column + 1):
                            cell = ws_din.cell(r, c)
                            cell.alignment = center_align
                            if r > 1 and c in [3, 6] and cell.value != '':
                                cell.number_format = '"R$" #,##0.00'

                # PLANILHA 3: RESUMO FINANCEIRO POR BANDEIRA
                if not df_macro_resumo.empty:
                    df_macro_resumo.to_excel(writer, index=False, sheet_name='Resumo por Bandeira')
                    ws_mac = writer.sheets['Resumo por Bandeira']
                    
                    for column in ws_mac.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                            except: pass
                        ws_mac.column_dimensions[col_letter].width = max(max_length + 4, 18)
                        
                    for r in range(1, ws_mac.max_row + 1):
                        for c in range(1, ws_mac.max_column + 1):
                            cell = ws_mac.cell(r, c)
                            cell.alignment = center_align
                            if r > 1 and c in [2, 3, 4]:
                                cell.number_format = '"R$" #,##0.00'

            st.download_button("📥 BAIXAR RESULTADO (.xlsx)", output.getvalue(), "conciliacao_pro.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("💡 Dica: Arraste os arquivos acima para começar.")
